import boto3
import pandas as pd
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime

# AWS Configuration
REGION = "namaregion"  # Change this to your AWS region
S3_BUCKET = "namabucket"  # Replace with your S3 bucket name

# AWS SES Configuration
SES_REGION = "ap-southeast-1"
SES_SENDER_EMAIL = "emailpengirim@example.com"
SES_RECIPIENT_EMAILS = ["emailpenerima1@example.com", "emailpenerima2@example.com"]  # Add more recipients

def get_ssm_managed_instances():
    """Retrieve a list of EC2 instances managed by AWS SSM."""
    ssm_client = boto3.client('ssm', region_name=REGION)
    response = ssm_client.describe_instance_information()
    return [instance['InstanceId'] for instance in response.get('InstanceInformationList', [])]

def get_ec2_instance_name(instance_id):
    """Retrieve the EC2 instance name from AWS EC2."""
    ec2_client = boto3.client('ec2', region_name=REGION)
    
    if not instance_id.startswith("i-"):
        return "N/A"

    response = ec2_client.describe_instances(InstanceIds=[instance_id])
    reservations = response.get('Reservations', [])
    
    if reservations:
        tags = reservations[0]['Instances'][0].get('Tags', [])
        for tag in tags:
            if tag['Key'] == 'Name':
                return tag['Value']
    
    return "N/A"

def format_cve_ids(cve_list):
    """Format CVE IDs by removing commas and spaces."""
    return "".join(cve_list) if cve_list else "N/A"

def get_missing_patches(instance_id):
    """Retrieve missing patches for a given EC2 instance."""
    ssm_client = boto3.client('ssm', region_name=REGION)
    response = ssm_client.describe_instance_patches(
        InstanceId=instance_id,
        Filters=[{'Key': 'State', 'Values': ['Missing']}]
    )

    patches = response.get('Patches', [])
    missing_patches = []

    for patch in patches:
        missing_patches.append((
            patch.get('KBId', 'N/A'),
            patch.get('Title', 'N/A'),
            patch.get('Classification', 'N/A'),
            patch.get('Severity', 'N/A'),
            patch.get('State', 'N/A'),
            format_cve_ids(patch.get('CVEIds', []))  # ✅ CVE formatting fix
        ))
    
    return missing_patches

def generate_excel(data):
    """Generate an Excel file with patch report data."""
    df = pd.DataFrame(data, columns=[
        "Instance ID", "Instance Name", "Missing Patch", 
        "Title", "Classification", "Severity", "State", "CVE IDs"
    ])
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Patch Report")
        ws = writer.book.active

        # Style header row
        header_fill = PatternFill(start_color="538DD5", end_color="538DD5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
    
    output.seek(0)  # Reset buffer position
    return output

def upload_to_s3(file_content, file_name):
    """Upload the generated Excel report to S3."""
    s3_client = boto3.client('s3')
    file_path = f"report/{file_name}"  # Store in 'report/' folder
    
    s3_client.put_object(
        Bucket=S3_BUCKET, 
        Key=file_path, 
        Body=file_content.getvalue(),
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return f"s3://{S3_BUCKET}/{file_path}", file_path  # Return S3 URL and path

def send_ses_email_with_attachment(bucket_name, file_path, file_name):
    """Send an email via AWS SES with the Excel report attached."""
    ses_client = boto3.client('ses', region_name=SES_REGION)
    s3_client = boto3.client('s3')

    try:
        # Retrieve the file from S3
        file_obj = s3_client.get_object(Bucket=bucket_name, Key=file_path)
        file_content = file_obj['Body'].read()
        encoded_file = base64.b64encode(file_content).decode('utf-8')

        # Email subject and body
        subject = "SSM Patch Report with Attachment"
        body_text = """Hello,

The latest SSM Patch Report has been generated.

The report is attached.

Regards,
Automated Patch Report System
"""

        # Email headers for multiple recipients and attachment
        email_message = f"""From: {SES_SENDER_EMAIL}
To: {", ".join(SES_RECIPIENT_EMAILS)}
Subject: {subject}
MIME-Version: 1.0
Content-Type: multipart/mixed; boundary="boundary-string"

--boundary-string
Content-Type: text/plain; charset=UTF-8
Content-Transfer-Encoding: 7bit

{body_text}

--boundary-string
Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; name="{file_name}"
Content-Transfer-Encoding: base64
Content-Disposition: attachment; filename="{file_name}"

{encoded_file}
--boundary-string--
"""

        # Send email
        response = ses_client.send_raw_email(
            Source=SES_SENDER_EMAIL,
            Destinations=SES_RECIPIENT_EMAILS,
            RawMessage={"Data": email_message}
        )

        return {"message": "Email sent successfully", "message_id": response["MessageId"]}

    except Exception as e:
        return {"message": "Failed to send email", "error": str(e)}

def lambda_handler(event, context):
    """AWS Lambda handler function."""
    
    instances = get_ssm_managed_instances()
    
    if not instances:
        return {"message": "No instances found in SSM."}
    
    data = []
    
    for instance_id in instances:
        instance_name = get_ec2_instance_name(instance_id)
        missing_patches = get_missing_patches(instance_id)

        if missing_patches:
            for kb_id, title, classification, severity, state, cve_ids in missing_patches:
                data.append([instance_id, instance_name, kb_id, title, classification, severity, state, cve_ids])
        else:
            data.append([instance_id, instance_name, "No missing patches", "N/A", "N/A", "N/A", "N/A", "No CVEs"])
    
    current_date = datetime.now().strftime("%Y-%m-%d")
    file_name = f"ssm_patch_report_{current_date}.xlsx"

    excel_data = generate_excel(data)
    file_url, file_path = upload_to_s3(excel_data, file_name)

    email_status = send_ses_email_with_attachment(S3_BUCKET, file_path, file_name)

    return {
        "message": "Report generated, uploaded, and email sent successfully.",
        "file_url": file_url,
        "email_status": email_status
    }
